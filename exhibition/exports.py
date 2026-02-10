from urllib.parse import quote
from io import BytesIO

from django.contrib import admin
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.urls import path, reverse
from django.db import models

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from exhibition.models import Exhibitions, Portfolio
from rating.models import Rating


class ExportExhibitionAdmin(admin.ModelAdmin):
	DEFAULT_PROJECTS_PER_NOMINATION = 3

	def get_urls(self):
		urls = super().get_urls()
		info = self.model._meta.app_label, self.model._meta.model_name

		custom_urls = [
			path(
				'<path:object_id>/export-jury-ratings/',
				self.admin_site.admin_view(self.export_jury_ratings),
				name='%s_%s_export_jury_ratings' % info,
			),
		]
		return custom_urls + urls

	def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
		extra_context = extra_context or {}
		if object_id:
			extra_context['show_export_button'] = True
			info = self.model._meta.app_label, self.model._meta.model_name
			extra_context['export_url'] = reverse('admin:%s_%s_export_jury_ratings' % info, args=[object_id])
		return super().changeform_view(request, object_id, form_url, extra_context)

	@staticmethod
	def _get_ratings_map(exhibition):
		ratings = (
			Rating.objects
			.filter(
				portfolio__exhibition=exhibition,
				is_jury_rating=True
			)
			.values('portfolio_id', 'user_id')
			.annotate(score=models.Sum('star'))
		)

		return {
			(r['portfolio_id'], r['user_id']): r['score']
			for r in ratings
		}

	def export_jury_ratings(self, request, object_id):
		"""Страница экспорта оценок жюри"""
		exhibition = get_object_or_404(Exhibitions, pk=object_id)

		projects_per_nom = request.GET.get('limit', self.DEFAULT_PROJECTS_PER_NOMINATION)
		try:
			projects_per_nom = min(int(projects_per_nom), 10)
		except (TypeError, ValueError):
			projects_per_nom = min(self.DEFAULT_PROJECTS_PER_NOMINATION, 10)

		# --- POST: экспорт ---
		if request.method == 'POST':
			filename = request.POST.get('filename', '').strip()
			if not filename:
				filename = f"jury_ratings_{exhibition.slug}"

			# Получаем те же данные, что и для HTML
			report_data = self._get_report_data(exhibition, projects_per_nom)

			wb = self._generate_excel_report(report_data)

			output = BytesIO()
			wb.save(output)
			output.seek(0)

			if not filename.lower().endswith('.xlsx'):
				filename += '.xlsx'

			encoded_filename = quote(filename)

			response = HttpResponse(
				output.read(),
				content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
			)
			response['Content-Disposition'] = (
				f"attachment; filename*=UTF-8''{encoded_filename}"
			)

			return response

		# --- GET: отображение страницы ---
		report_data = self._get_report_data(exhibition, projects_per_nom)

		context = {
			**self.admin_site.each_context(request),
			'title': f'Протокол оценок жюри: {exhibition.title}',
			'exhibition': exhibition,
			'report_data': report_data,
			'projects_per_nomination': projects_per_nom,
			'opts': self.model._meta,
		}

		return render(
			request,
			'admin/exhibition/jury_ratings_export.html',
			context
		)

	def _get_report_data(self, exhibition, projects_per_nomination):
		jury_list = list(exhibition.jury.select_related('user'))
		ratings_map = self._get_ratings_map(exhibition)

		portfolios = (
			Portfolio.objects
			.filter(exhibition=exhibition, status=True)
			.prefetch_related('nominations')
			.select_related('owner')
		)

		portfolios_by_nom = {}
		for p in portfolios:
			for nom in p.nominations.all():
				portfolios_by_nom.setdefault(nom.id, []).append(p)

		report_data = {
			'title': f'Протокол оценок жюри: {exhibition.title}',
			'jury_list': jury_list,
			'nominations': [],
			'jury_stats': {},
			'not_voted_jury': {},
			'total_stats': {
				'total_jury': len(jury_list),
				'total_nominations': exhibition.nominations.count(),
				'total_projects': 0,
				'total_ratings': 0,
			}
		}

		# jury stats
		for jury in jury_list:
			count = Rating.objects.filter(
				user=jury.user,
				is_jury_rating=True,
				portfolio__exhibition=exhibition
			).count()

			report_data['jury_stats'][jury.id] = {'name': jury.name or jury.user_name}
			report_data['total_stats']['total_ratings'] += count
			report_data['not_voted_jury'][jury.id] = {
				'jury': jury,
				'nominations': [],
				'total_missing': 0
			}

		for nomination in exhibition.nominations.all():
			projects_data = []
			jury_vote_counter = {j.id: 0 for j in jury_list}
			portfolios = portfolios_by_nom.get(nomination.id, [])

			for portfolio in portfolios:
				jury_scores = {}

				for jury in jury_list:
					score = ratings_map.get((portfolio.id, jury.user.id))
					jury_scores[jury.id] = score
					if score is not None:
						jury_vote_counter[jury.id] += 1

				stats = portfolio.get_rating_stats()

				projects_data.append({
					'portfolio': portfolio,
					'jury_scores': jury_scores,  # ← оставляем, если нужен Excel поимённо
					'total_score': stats['jury_average'],
					'votes': stats['jury_count'],
				})

			# победители (логика сохранена)
			winners = []
			total_possible = len(jury_list) * len(projects_data)
			actual_votes = sum(jury_vote_counter.values())

			if projects_data and actual_votes == total_possible:
				max_score = max(p['total_score'] for p in projects_data)
				winners = [{
					'portfolio': p['portfolio'],
					'score': p['total_score'],
					'medal': 'gold',
					'position': 1
				} for p in projects_data if p['total_score'] == max_score]

			# не проголосовавшие
			for jury in jury_list:
				expected = len(projects_data)
				actual = jury_vote_counter[jury.id]
				if actual < expected:
					report_data['not_voted_jury'][jury.id]['nominations'].append({
						'nomination': nomination,
						'voted': actual,
						'total': expected,
						'missing': expected - actual
					})
					report_data['not_voted_jury'][jury.id]['total_missing'] += expected - actual

			projects_data.sort(key=lambda x: x['total_score'], reverse=True)

			report_data['nominations'].append({
				'title': nomination.title,
				'all_projects': len(projects_data),
				'top_projects': projects_data[:projects_per_nomination],
				'other_projects': projects_data[projects_per_nomination:],
				'winners': winners,
				'jury_counts': jury_vote_counter,
			})

			report_data['total_stats']['total_projects'] += len(projects_data)

		report_data['not_voted_jury'] = {
			k: v for k, v in report_data['not_voted_jury'].items()
			if v['total_missing'] > 0
		}

		return report_data

	@staticmethod
	def _add_jury_control_sheet(wb, report_data):
		ws = wb.create_sheet(title='Контроль жюри')

		bold = Font(bold=True)
		center = Alignment(horizontal='center', vertical='center')
		left = Alignment(horizontal='left', vertical='center')

		row = 1

		# --- Заголовок документа ---
		ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
		ws.cell(row=row, column=1, value='КОНТРОЛЬ ГОЛОСОВАНИЯ ЖЮРИ')
		ws.cell(row=row, column=1).font = Font(bold=True, size=14)
		ws.cell(row=row, column=1).alignment = center
		row += 2

		# --- Header таблицы ---
		headers = ['Жюри', 'Номинация', 'Проголосовано', 'Должно', 'Пропущено']
		for col, title in enumerate(headers, start=1):
			ws.cell(row=row, column=col, value=title).font = bold
			ws.cell(row=row, column=col).alignment = center

		row += 1

		for item in report_data['not_voted_jury'].values():
			jury_name = item['jury'].name or item['jury'].user_name

			for nom in item['nominations']:
				ws.cell(row=row, column=1, value=jury_name).alignment = left
				ws.cell(row=row, column=2, value=nom['nomination'].title).alignment = left
				ws.cell(row=row, column=3, value=nom['voted']).alignment = center
				ws.cell(row=row, column=4, value=nom['total']).alignment = center
				ws.cell(row=row, column=5, value=nom['missing']).alignment = center

				row += 1

		ws.column_dimensions['A'].width = 30
		ws.column_dimensions['B'].width = 35
		ws.column_dimensions['C'].width = 14
		ws.column_dimensions['D'].width = 14
		ws.column_dimensions['E'].width = 14

	@staticmethod
	def _add_winners_summary_sheet( wb, report_data):
		ws = wb.create_sheet(title='Сводка победителей')

		bold = Font(bold=True)
		center = Alignment(horizontal='center', vertical='center')
		left = Alignment(horizontal='left', vertical='center')

		row = 1

		# --- Заголовок документа ---
		ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
		ws.cell(row=row, column=1, value='СВОДКА ПОБЕДИТЕЛЕЙ')
		ws.cell(row=row, column=1).font = Font(bold=True, size=15)
		ws.cell(row=row, column=1).alignment = center
		row += 2

		# --- Header таблицы ---
		headers = ['Номинация', 'Проект', 'Баллы']
		for col, title in enumerate(headers, start=1):
			ws.cell(row=row, column=col, value=title).font = bold
			ws.cell(row=row, column=col).alignment = center

		row += 1

		for nomination in report_data['nominations']:
			for winner in nomination['winners']:
				ws.cell(row=row, column=1, value=nomination['title']).alignment = left
				ws.cell(row=row, column=2, value=str(winner['portfolio'])).alignment = left
				ws.cell(row=row, column=3, value=winner['score']).alignment = center

				row += 1

		# ширины
		ws.column_dimensions['A'].width = 60
		ws.column_dimensions['B'].width = 40
		ws.column_dimensions['C'].width = 10

	def _generate_excel_report(self, report_data):
		wb = Workbook()
		ws = wb.active
		ws.title = 'Итоги'
		title = report_data['title'] or 'ИТОГОВЫЙ ПРОТОКОЛ ЖЮРИ'

		jury_list = report_data['jury_list']

		first_jury_col = 2
		total_col = first_jury_col + len(jury_list)
		max_col = total_col + 1
		winner_col = max_col + 1
		last_table_col = winner_col

		bold = Font(bold=True)
		center = Alignment(horizontal='center', vertical='center', wrap_text=True)
		left = Alignment(horizontal='left', vertical='center', wrap_text=True)
		thin = Side(style='thin')
		border = Border(left=thin, right=thin, top=thin, bottom=thin)

		row = 1

		# ===== Заголовок документа =====
		ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_table_col)
		ws.cell(row=row, column=1, value=title)
		ws.cell(row=row, column=1).font = Font(bold=True, size=15)
		ws.cell(row=row, column=1).alignment = center
		row += 2

		# ===== По номинациям =====
		for nomination in report_data['nominations']:
			ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_table_col)
			ws.cell(row=row, column=1, value=f"Номинация: {nomination['title']}")
			ws.cell(row=row, column=1).font = Font(bold=True, size=13)
			ws.cell(row=row, column=1).alignment = left
			row += 1

			# ---- Header ----
			headers = ['Проект']
			headers += [jury.name or jury.user_name for jury in jury_list]
			headers.append('Итого')
			headers.append('Макс.')
			headers.append('Победитель')

			for col, title in enumerate(headers, start=1):
				cell = ws.cell(row=row, column=col, value=title)
				cell.font = bold
				cell.alignment = center
				cell.border = border

			row += 1
			start_projects_row = row

			# ---- Projects ----
			for project in nomination['top_projects'] + nomination['other_projects']:
				col = 1
				cell = ws.cell(row=row, column=col, value=str(project['portfolio']))
				cell.alignment = left
				cell.border = border

				score_cells = []
				col += 1

				for jury in jury_list:
					score = project['jury_scores'].get(jury.id)
					cell = ws.cell(row=row, column=col, value=score)
					cell.alignment = center
					cell.border = border
					score_cells.append(cell.coordinate)
					col += 1

				# SUM
				sum_cell = ws.cell(
					row=row,
					column=col,
					value=f"=SUM({score_cells[0]}:{score_cells[-1]})"
				)
				sum_cell.alignment = center
				sum_cell.border = border
				col += 1

				row += 1

			end_projects_row = row - 1

			# ---- MAX column ----
			total_col = len(jury_list) + 2
			max_col = total_col + 1
			winner_col = max_col + 1

			for r in range(start_projects_row, end_projects_row + 1):
				ws.cell(
					row=r,
					column=max_col,
					value=f"=MAX(${get_column_letter(total_col)}${start_projects_row}:${get_column_letter(total_col)}${end_projects_row})"
				).border = border

				cell = ws.cell(
					row=r,
					column=winner_col,
					value=f'=IF({get_column_letter(total_col)}{r}={get_column_letter(max_col)}{r},"①","")'
				)
				cell.font = Font(bold=True)
				cell.alignment = center
				cell.border = border

			ws.column_dimensions[get_column_letter(max_col)].hidden = True
			# ---- Borders for MAX / Winner headers ----
			ws.cell(row=start_projects_row - 1, column=max_col).font = bold
			ws.cell(row=start_projects_row - 1, column=winner_col).font = bold
			ws.cell(row=start_projects_row - 1, column=max_col).alignment = center
			ws.cell(row=start_projects_row - 1, column=winner_col).alignment = center

			row += 2

		# ===== Автоширина =====
		for col in range(1, ws.max_column + 1):
			ws.column_dimensions[get_column_letter(col)].width = 18

		self._add_winners_summary_sheet(wb, report_data)
		self._add_jury_control_sheet(wb, report_data)

		return wb

